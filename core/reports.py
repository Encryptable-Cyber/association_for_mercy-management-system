import io
import json
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import openpyxl


def log_export(user, report_type, export_type, filters=None, ip_address=None):
    """Record an export in the audit log."""
    from .models import ExportLog
    ExportLog.objects.create(
        user=user,
        report_type=report_type,
        export_type=export_type,
        filters_applied=json.dumps(filters) if filters else '',
        ip_address=ip_address
    )


def generate_pdf_report(title, headers, data, filename='report.pdf'):
    """
    Generate a professional PDF report.
    Returns an HttpResponse with PDF content.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=15*mm,
        bottomMargin=15*mm,
        leftMargin=10*mm,
        rightMargin=10*mm,
        title=title
    )

    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=6,
        textColor=colors.HexColor('#1e3a5f')
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )
    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.black
    )

    # Title
    elements.append(Paragraph(f"Association for Mercy", title_style))
    elements.append(Paragraph(title, styles['Heading2']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 5*mm))

    # Table
    table_data = [
        [Paragraph(h, header_style) for h in headers]
    ]
    for row in data:
        table_data.append([Paragraph(str(cell), cell_style) for cell in row])

    col_width = (doc.width - 20) / len(headers)
    table = Table(table_data, colWidths=[col_width] * len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_excel_report(headers, data, filename='report.xlsx'):
    """
    Generate an Excel report.
    Returns an HttpResponse with Excel content.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Header styling
    header_fill = openpyxl.styles.PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    header_font = openpyxl.styles.Font(color='FFFFFF', bold=True, size=10)
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    # Write title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"Association for Mercy - {filename.replace('.xlsx', '')}")
    title_cell.font = openpyxl.styles.Font(bold=True, size=14, color='1e3a5f')
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y')}")
    ws.cell(row=2, column=1).font = openpyxl.styles.Font(italic=True, size=9, color='666666')

    # Write headers (row 4)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    # Write data (starting row 5)
    for row_idx, row in enumerate(data, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '—')
            cell.border = thin_border
            cell.font = openpyxl.styles.Font(size=9)

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response